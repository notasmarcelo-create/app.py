import streamlit as st
import pandas as pd
import datetime
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import (
    init_db, get_todas_movimentacoes, get_resumo_totais, 
    get_saldo_por_categoria, get_evolucao_mensal, get_resumo_por_produtores,
    add_movimentacao, update_movimentacao, delete_movimentacao, get_movimentacao,
    get_categorias_cadastradas, add_categoria_personalizada, delete_categoria_personalizada,
    get_produtores, add_produtor, delete_produtor,
    autenticar_usuario, alterar_senha_usuario, get_usuarios_cadastrados, add_novo_usuario,
    TIPOS_LANCAMENTO_PADRAO, formatar_data_para_exibicao
)
from seed_data import popular_banco

# Configuração da Página
st.set_page_config(
    page_title="Controle de Rebanho - Marcelo & Michele",
    page_icon="🐄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inicializa o Banco
init_db()

# Inicializa Estado de Autenticação e Importação
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "usuario_atual" not in st.session_state:
    st.session_state.usuario_atual = None
if "msg_sucesso_importacao" not in st.session_state:
    st.session_state.msg_sucesso_importacao = None

# ============================================================
# FUNÇÕES DE EXPORTAÇÃO E IMPORTAÇÃO INTELIGENTE (EMBUTIDAS)
# ============================================================
def gerar_relatorio_excel_bytes(filtros=None):
    output = io.BytesIO()
    produtor_filtro = filtros.get("produtor") if filtros else None
    df_mov = get_todas_movimentacoes(filtros)
    df_saldo = get_saldo_por_categoria(produtor_filtro)
    resumo = get_resumo_totais(produtor_filtro)
    df_resumo_prods = get_resumo_por_produtores()

    if not df_mov.empty:
        colunas_export = ["PRODUTOR", "LANÇAMENTO", "DATA", "OPERAÇÃO", "QTD. TOTAL", "GTA", "NFP", "ANIMAIS / MESES", "OBSERVAÇÕES"]
        df_export = df_mov[colunas_export].copy()
    else:
        df_export = pd.DataFrame(columns=["PRODUTOR", "LANÇAMENTO", "DATA", "OPERAÇÃO", "QTD. TOTAL", "GTA", "NFP", "ANIMAIS / MESES", "OBSERVAÇÕES"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Lançamentos", index=False)
        df_saldo.to_excel(writer, sheet_name="Saldo por Categoria", index=False)
        df_resumo_prods.to_excel(writer, sheet_name="Resumo por Produtor", index=False)

        prod_label = produtor_filtro if produtor_filtro and produtor_filtro not in ("TODOS", "Todos os Produtores") else "Consolidado (Todos)"
        df_resumo = pd.DataFrame([
            {"Indicador": "Produtor Selecionado", "Valor": prod_label},
            {"Indicador": "Saldo Atual do Rebanho (Cabeças)", "Valor": resumo["saldo_atual"]},
            {"Indicador": "Total de Entradas (Créditos)", "Valor": resumo["total_entradas"]},
            {"Indicador": "Total de Saídas (Débitos)", "Valor": resumo["total_saidas"]},
            {"Indicador": "Total de Movimentações Registradas", "Valor": resumo["total_registros"]}
        ])
        df_resumo.to_excel(writer, sheet_name="Resumo Geral", index=False)

        workbook = writer.book
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        data_font = Font(name="Arial", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for sheet in workbook.worksheets:
            sheet.views.sheetView[0].showGridLines = True
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0"
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output.seek(0)
    return output.getvalue()

def extrair_detalhes_de_texto(texto_categorias, qtd_total=0):
    if not texto_categorias or pd.isna(texto_categorias) or not str(texto_categorias).strip():
        return [("DIVERSOS / OUTROS", qtd_total)]
    texto = str(texto_categorias).strip()
    padrao = r'(\d+)\s*[-–]\s*([A-Za-zÀ-ÿ0-9\+\s]+?)(?=(?:\s+\d+\s*[-–]|$|\|))'
    matches = re.findall(padrao, texto)
    if matches:
        detalhes = []
        for qtd_str, cat_str in matches:
            detalhes.append((cat_str.strip().upper(), int(qtd_str)))
        return detalhes
    else:
        return [(texto.upper(), qtd_total)]

def carregar_dataframe_inteligente(arquivo_bytes_ou_upload):
    if hasattr(arquivo_bytes_ou_upload, "getvalue"):
        conteudo_bytes = arquivo_bytes_ou_upload.getvalue()
    elif hasattr(arquivo_bytes_ou_upload, "read"):
        arquivo_bytes_ou_upload.seek(0)
        conteudo_bytes = arquivo_bytes_ou_upload.read()
    else:
        conteudo_bytes = arquivo_bytes_ou_upload

    df_raw = None
    try:
        excel_file = pd.ExcelFile(io.BytesIO(conteudo_bytes))
        for sheet_name in excel_file.sheet_names:
            df_temp = pd.read_excel(io.BytesIO(conteudo_bytes), sheet_name=sheet_name, header=None)
            if not df_temp.empty and len(df_temp) > 1:
                df_raw = df_temp
                break
        if df_raw is None:
            df_raw = pd.read_excel(io.BytesIO(conteudo_bytes), header=None)
    except Exception:
        try:
            df_raw = pd.read_csv(io.BytesIO(conteudo_bytes), header=None, encoding="utf-8")
        except Exception:
            try:
                df_raw = pd.read_csv(io.BytesIO(conteudo_bytes), header=None, encoding="latin1", sep=";")
            except Exception as e:
                raise ValueError(f"Não foi possível ler o arquivo: {str(e)}")

    if df_raw is None or df_raw.empty:
        raise ValueError("O arquivo enviado está vazio ou ilegível.")

    header_idx = None
    for r_idx in range(min(20, len(df_raw))):
        row_values = [str(val).strip().upper() for val in df_raw.iloc[r_idx].dropna()]
        tem_lanc = any("LANÇ" in v or "LANC" in v or "TIPO" in v or "OPER" in v for v in row_values)
        tem_data = any("DATA" in v or "DARA" in v or "DIA" in v for v in row_values)
        tem_qtd = any("QTD" in v or "QUANT" in v or "TOTAL" in v or "CABEC" in v for v in row_values)
        if (tem_lanc and tem_data) or (tem_data and tem_qtd) or (tem_lanc and tem_qtd):
            header_idx = r_idx
            break

    if header_idx is not None:
        headers = [str(val).strip() if pd.notna(val) else f"COL_{i}" for i, val in enumerate(df_raw.iloc[header_idx])]
        df = df_raw.iloc[header_idx + 1:].copy()
        df.columns = headers
        df = df.reset_index(drop=True)
    else:
        headers = [str(val).strip() if pd.notna(val) else f"COL_{i}" for i, val in enumerate(df_raw.iloc[0])]
        df = df_raw.iloc[1:].copy()
        df.columns = headers
        df = df.reset_index(drop=True)

    col_map = {}
    for col in df.columns:
        c_upper = str(col).strip().upper()
        if "PROD" in c_upper:
            col_map[col] = "PRODUTOR"
        elif "LANÇ" in c_upper or "LANC" in c_upper or "TIPO" in c_upper or "MOVIMENT" in c_upper:
            col_map[col] = "LANÇAMENTO"
        elif "DATA" in c_upper or "DARA" in c_upper or "DIA" in c_upper:
            col_map[col] = "DATA"
        elif "OPER" in c_upper or "C/D" in c_upper or "TIPO OP" in c_upper:
            col_map[col] = "OPERAÇÃO"
        elif "QTD" in c_upper or "QUANT" in c_upper or "TOTAL" in c_upper or "CABEC" in c_upper:
            col_map[col] = "QTD. TOTAL"
        elif "GTA" in c_upper or "GUIA" in c_upper:
            col_map[col] = "GTA"
        elif "NFP" in c_upper or "NOTA" in c_upper or "NF" in c_upper or "DOC" in c_upper:
            col_map[col] = "NFP"
        elif "ANIMA" in c_upper or "CATEG" in c_upper or "MESES" in c_upper or "IDADE" in c_upper:
            col_map[col] = "ANIMAIS / MESES"
        elif "OBS" in c_upper:
            col_map[col] = "OBSERVAÇÕES"

    df = df.rename(columns=col_map)
    df = df.dropna(how="all").reset_index(drop=True)
    return df

def importar_planilha_excel(arquivo_bytes_ou_upload, produtor_padrao="Michele"):
    try:
        df = carregar_dataframe_inteligente(arquivo_bytes_ou_upload)
    except Exception as e:
        return 0, [str(e)]

    if "LANÇAMENTO" not in df.columns:
        if "OPERAÇÃO" in df.columns:
            df["LANÇAMENTO"] = df["OPERAÇÃO"].apply(lambda op: "ENTRADA P/ CRIA / ENGORDA" if "CR" in str(op).upper() else "SAÍDA P/ CRIA / ENGORDA")
        else:
            df["LANÇAMENTO"] = "ENTRADA P/ CRIA / ENGORDA"

    if "OPERAÇÃO" not in df.columns:
        df["OPERAÇÃO"] = df["LANÇAMENTO"].apply(lambda l: "CRÉDITO" if "ENTRADA" in str(l).upper() or "NASC" in str(l).upper() or "COMPRA" in str(l).upper() else "DÉBITO")

    if "DATA" not in df.columns:
        df["DATA"] = datetime.date.today().strftime("%d/%m/%Y")

    if "QTD. TOTAL" not in df.columns:
        return 0, ["Coluna de Quantidade (QTD. TOTAL) não encontrada na planilha."]

    total_importados = 0
    erros = []

    for index, row in df.iterrows():
        try:
            tipo_lancamento = str(row["LANÇAMENTO"]).strip()
            if pd.isna(tipo_lancamento) or not tipo_lancamento or tipo_lancamento.lower() in ["nan", "none", "null", "total", "saldo"]:
                continue

            produtor = str(row["PRODUTOR"]).strip() if "PRODUTOR" in df.columns and pd.notna(row["PRODUTOR"]) and str(row["PRODUTOR"]).lower() not in ["nan", "none"] else produtor_padrao

            raw_data = row["DATA"]
            if isinstance(raw_data, (pd.Timestamp, pd.DatetimeIndex)):
                data_str = raw_data.strftime("%d/%m/%Y")
            elif hasattr(raw_data, "strftime"):
                data_str = raw_data.strftime("%d/%m/%Y")
            else:
                data_str = str(raw_data).strip()
                if data_str.lower() in ["nan", "none", ""]:
                    data_str = datetime.date.today().strftime("%d/%m/%Y")

            operacao = str(row["OPERAÇÃO"]).strip().upper()
            if "CR" in operacao or "ENTRADA" in operacao or "+" in operacao or "COMPRA" in operacao or "NASC" in operacao:
                operacao = "CRÉDITO"
            else:
                operacao = "DÉBITO"

            try:
                val_limpo = str(row["QTD. TOTAL"]).replace(".", "").replace(",", ".").replace(" ", "")
                num_matches = re.findall(r'\d+', val_limpo)
                qtd_total = int(num_matches[0]) if num_matches else 0
            except Exception:
                qtd_total = 0

            if qtd_total <= 0:
                continue

            gta = str(row["GTA"]).strip() if "GTA" in df.columns and pd.notna(row["GTA"]) and str(row["GTA"]).lower() not in ["nan", "none"] else ""
            nfp = str(row["NFP"]).strip() if "NFP" in df.columns and pd.notna(row["NFP"]) and str(row["NFP"]).lower() not in ["nan", "none"] else ""
            obs = str(row["OBSERVAÇÕES"]).strip() if "OBSERVAÇÕES" in df.columns and pd.notna(row["OBSERVAÇÕES"]) and str(row["OBSERVAÇÕES"]).lower() not in ["nan", "none"] else ""
            animais_texto = str(row["ANIMAIS / MESES"]).strip() if "ANIMAIS / MESES" in df.columns and pd.notna(row["ANIMAIS / MESES"]) and str(row["ANIMAIS / MESES"]).lower() not in ["nan", "none"] else ""

            detalhes = extrair_detalhes_de_texto(animais_texto, qtd_total)

            add_movimentacao(
                produtor=produtor,
                tipo_lancamento=tipo_lancamento,
                data=data_str,
                operacao=operacao,
                qtd_total=qtd_total,
                gta=gta,
                nfp=nfp,
                detalhes=detalhes,
                observacoes=obs
            )
            total_importados += 1
        except Exception as e:
            erros.append(f"Linha {index + 1}: {str(e)}")

    return total_importados, erros


# Estilos CSS Personalizados
st.markdown("""
<style>
    .main-header {
        font-size: 26px;
        font-weight: 700;
        color: #1b4332;
        margin-bottom: 2px;
    }
    .sub-header {
        font-size: 14px;
        color: #555;
        margin-bottom: 15px;
    }
    .producer-badge {
        display: inline-block;
        padding: 5px 14px;
        border-radius: 20px;
        font-size: 13px;
        font-weight: 600;
        background-color: #e8f5e9;
        color: #2e7d32;
        border: 1px solid #c8e6c9;
        margin-bottom: 10px;
    }
    .metric-card {
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
        border-radius: 10px;
        padding: 16px;
        border-left: 5px solid #2d6a4f;
        box-shadow: 0 2px 6px rgba(0,0,0,0.06);
    }
    .metric-title {
        font-size: 12px;
        font-weight: 600;
        color: #495057;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .metric-value {
        font-size: 26px;
        font-weight: 700;
        color: #1b4332;
        margin-top: 4px;
    }
    .prod-card {
        background: #f1f8f5;
        border: 1px solid #d8ebd4;
        border-radius: 8px;
        padding: 12px 16px;
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# TELA DE LOGIN (Se não estiver autenticado)
# ==========================================
if not st.session_state.autenticado:
    col_l1, col_l2, col_l3 = st.columns([1, 1.4, 1])
    with col_l2:
        st.write("")
        st.write("")
        st.markdown("""
        <div style="text-align: center; margin-bottom: 20px;">
            <div style="font-size: 48px;">🐄</div>
            <h2 style="color: #1b4332; margin-top: 5px; font-weight: 700;">Gestão de Rebanho</h2>
            <p style="color: #666; font-size: 14px;">Lançamentos de Gado • Marcelo & Michele</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("form_login"):
            st.markdown("#### 🔒 Acesso ao Sistema")
            username_input = st.text_input("Usuário", placeholder="michele, marcelo ou admin")
            senha_input = st.text_input("Senha", type="password", placeholder="Digite sua senha")
            
            btn_entrar = st.form_submit_button("🚀 Entrar no Sistema", type="primary")

            if btn_entrar:
                if username_input.strip() and senha_input.strip():
                    usuario_valido = autenticar_usuario(username_input, senha_input)
                    if usuario_valido:
                        st.session_state.autenticado = True
                        st.session_state.usuario_atual = usuario_valido
                        if usuario_valido["produtor_associado"] in ["Michele", "Marcelo"]:
                            st.session_state.produtor_ativo = usuario_valido["produtor_associado"]
                        else:
                            st.session_state.produtor_ativo = "Todos os Produtores"
                        st.success(f"Bem-vindo(a), {usuario_valido['nome_completo']}!")
                        st.rerun()
                    else:
                        st.error("❌ Usuário ou senha incorretos. Tente novamente.")
                else:
                    st.warning("Preencha usuário e senha.")

        with st.expander("ℹ️ Acessos Padrão Configurados"):
            st.markdown("""
            | Usuário | Senha Padrão | Produtor Associado |
            | :--- | :--- | :--- |
            | **michele** | `michele123` | Michele |
            | **marcelo** | `marcelo123` | Marcelo |
            | **admin** | `admin123` | Todos os Produtores |
            
            *(Você poderá alterar a senha depois de entrar no sistema)*
            """)
    st.stop()

# ==========================================
# SISTEMA PRINCIPAL (APÓS LOGIN)
# ==========================================

usuario_logado = st.session_state.usuario_atual
lista_produtores = get_produtores()
if not lista_produtores:
    lista_produtores = ["Michele", "Marcelo"]

# Sidebar
with st.sidebar:
    st.markdown(f"### 👤 Olá, {usuario_logado['nome_completo']}!")
    st.caption(f"Usuário: `{usuario_logado['username']}`")
    
    if st.button("🚪 Sair / Desconectar"):
        st.session_state.autenticado = False
        st.session_state.usuario_atual = None
        st.rerun()

    st.divider()

    st.markdown("#### 🚜 Visualização de Produtor")
    opcoes_produtor_global = ["Todos os Produtores"] + lista_produtores
    
    if "produtor_ativo" not in st.session_state:
        st.session_state.produtor_ativo = "Todos os Produtores"

    produtor_selecionado = st.selectbox(
        "Visualizar dados de:",
        opcoes_produtor_global,
        index=opcoes_produtor_global.index(st.session_state.produtor_ativo) if st.session_state.produtor_ativo in opcoes_produtor_global else 0,
        key="select_produtor_global"
    )
    st.session_state.produtor_ativo = produtor_selecionado

    st.markdown("---")
    with st.expander("🔑 Alterar Minha Senha"):
        with st.form("form_troca_senha_sidebar"):
            nova_pwd = st.text_input("Nova Senha", type="password")
            conf_pwd = st.text_input("Confirmar Nova Senha", type="password")
            btn_trocar = st.form_submit_button("Salvar Nova Senha", type="primary")
            if btn_trocar:
                if nova_pwd and nova_pwd == conf_pwd:
                    alterar_senha_usuario(usuario_logado["username"], nova_pwd)
                    st.success("Senha alterada com sucesso!")
                else:
                    st.error("As senhas não coincidem ou estão vazias.")

    st.caption("Sistema de Gestão Agropecuária v2.0 • Protegido com Login")

# Top Bar
col_head1, col_head2 = st.columns([3, 1])
with col_head1:
    st.markdown('<div class="main-header">🐄 Sistema de Gestão de Rebanho & Lançamentos</div>', unsafe_allow_html=True)
    if produtor_selecionado != "Todos os Produtores":
        st.markdown(f'<div class="producer-badge">👤 Filtrando por: <b>{produtor_selecionado}</b></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="producer-badge" style="background-color: #e3f2fd; color: #1565c0; border-color: #bbdefb;">👥 <b>Visão Consolidada (Michele e Marcelo)</b></div>', unsafe_allow_html=True)
with col_head2:
    excel_data_quick = gerar_relatorio_excel_bytes({"produtor": produtor_selecionado})
    st.download_button(
        label="📥 Baixar Excel",
        data=excel_data_quick,
        file_name=f"rebanho_{produtor_selecionado.lower().replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Abas
tab_dashboard, tab_novo, tab_extrato, tab_import_export, tab_ajustes = st.tabs([
    "📊 Painel Geral & Estoque",
    "➕ Novo Lançamento",
    "📋 Histórico & Extrato",
    "📤 Importar / Exportar",
    "⚙️ Usuários & Configurações"
])

# ==========================================
# ABA 1: PAINEL GERAL & ESTOQUE
# ==========================================
with tab_dashboard:
    resumo = get_resumo_totais(produtor_selecionado)
    df_saldo_cat = get_saldo_por_categoria(produtor_selecionado)
    df_evolucao = get_evolucao_mensal(produtor_selecionado)
    df_resumo_prods = get_resumo_por_produtores()

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #2d6a4f;">
            <div class="metric-title">🐂 Saldo Atual do Rebanho</div>
            <div class="metric-value">{resumo['saldo_atual']:,} <span style="font-size: 13px; font-weight: normal; color: #666;">cabeças</span></div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #2b9348;">
            <div class="metric-title">📥 Total Entradas (Crédito)</div>
            <div class="metric-value" style="color: #2b9348;">+{resumo['total_entradas']:,}</div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #d90429;">
            <div class="metric-title">📤 Total Saídas (Débito)</div>
            <div class="metric-value" style="color: #d90429;">-{resumo['total_saidas']:,}</div>
        </div>
        """, unsafe_allow_html=True)
    with c4:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #0077b6;">
            <div class="metric-title">📄 Total Lançamentos</div>
            <div class="metric-value" style="color: #0077b6;">{resumo['total_registros']}</div>
        </div>
        """, unsafe_allow_html=True)

    st.write("")
    
    if produtor_selecionado == "Todos os Produtores" and not df_resumo_prods.empty:
        st.markdown("#### 👥 Resumo Comparativo por Produtor")
        cols_prods = st.columns(len(df_resumo_prods))
        for idx, row_p in df_resumo_prods.iterrows():
            with cols_prods[idx % len(cols_prods)]:
                st.markdown(f"""
                <div class="prod-card">
                    <div style="font-size: 15px; font-weight: bold; color: #1b4332;">👤 {row_p['Produtor']}</div>
                    <div style="font-size: 20px; font-weight: bold; color: #2d6a4f; margin: 4px 0;">{row_p['Saldo Atual']:,} <span style="font-size: 12px; font-weight: normal; color: #555;">cab. no estoque</span></div>
                    <div style="font-size: 12px; color: #555;">Entradas: <b>+{row_p['Entradas']}</b> | Saídas: <b>-{row_p['Saídas']}</b> ({row_p['Lançamentos']} docs)</div>
                </div>
                """, unsafe_allow_html=True)

    st.divider()

    col_cat, col_graf = st.columns([1.1, 1.2])

    with col_cat:
        st.subheader(f"📦 Saldo por Categoria / Idade ({produtor_selecionado})")
        if not df_saldo_cat.empty:
            st.dataframe(
                df_saldo_cat,
                column_config={
                    "Categoria": st.column_config.TextColumn("Categoria / Idade", width="medium"),
                    "Entradas (Crédito)": st.column_config.NumberColumn("Entradas", format="%d"),
                    "Saídas (Débito)": st.column_config.NumberColumn("Saídas", format="%d"),
                    "Saldo Atual": st.column_config.NumberColumn("Saldo Atual", format="%d")
                },
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("Nenhum saldo registrado para este produtor.")

    with col_graf:
        st.subheader(f"📈 Movimentações Mensais ({produtor_selecionado})")
        if not df_evolucao.empty:
            chart_data = df_evolucao.set_index("mes_ano")[["entradas", "saidas"]]
            chart_data.columns = ["Entradas (Crédito)", "Saídas (Débito)"]
            st.bar_chart(chart_data, color=["#2b9348", "#d90429"], height=320)
        else:
            st.info("Sem dados suficientes para gráficos.")

# ==========================================
# ABA 2: NOVO LANÇAMENTO
# ==========================================
with tab_novo:
    st.subheader("📝 Cadastrar Nova Movimentação")
    st.caption("Selecione o produtor rural (Marcelo ou Michele), dados do documento e detalhe os animais.")

    categorias_disponiveis = get_categorias_cadastradas()
    if not categorias_disponiveis:
        categorias_disponiveis = ["DIVERSOS / OUTROS", "VACAS +36", "NOVILHAS 13-24", "NOVILHAS 0-12", "TERNEIROS 0-12"]

    idx_prod_form = 0
    if produtor_selecionado in lista_produtores:
        idx_prod_form = lista_produtores.index(produtor_selecionado)

    if "num_linhas_categoria" not in st.session_state:
        st.session_state.num_linhas_categoria = 1

    with st.form("form_novo_lancamento", clear_on_submit=True):
        col_f0, col_f1, col_f2, col_f3 = st.columns([1.2, 1.5, 1, 1.2])
        with col_f0:
            prod_lancamento = st.selectbox("Produtor Rural *", lista_produtores, index=idx_prod_form)
        with col_f1:
            tipo_lanc = st.selectbox("Tipo de Lançamento *", TIPOS_LANCAMENTO_PADRAO)
        with col_f2:
            data_mov = st.date_input("Data *", value=datetime.date.today(), format="DD/MM/YYYY")
        with col_f3:
            operacao = st.selectbox(
                "Operação *",
                ["DÉBITO (Saída/Abate/Venda)", "CRÉDITO (Entrada/Compra/Nascimento)"]
            )

        col_f4, col_f5 = st.columns(2)
        with col_f4:
            gta = st.text_input("Número da GTA (Guia de Trânsito Animal)", placeholder="Ex: AD-616710")
        with col_f5:
            nfp = st.text_input("Número da NFP (Nota Fiscal de Produtor)", placeholder="Ex: 64117201")

        st.markdown("##### 🐄 Detalhamento dos Animais")
        st.caption("Informe as categorias e quantidades deste lote:")

        detalhes_form = []
        qtd_somada = 0

        for i in range(st.session_state.num_linhas_categoria):
            c_cat, c_qtd = st.columns([3, 1])
            with c_cat:
                cat_selecionada = st.selectbox(
                    f"Categoria / Faixa Etária #{i+1}",
                    categorias_disponiveis,
                    key=f"novo_cat_{i}"
                )
            with c_qtd:
                qtd_categoria = st.number_input(
                    f"Quantidade #{i+1}",
                    min_value=0,
                    value=0,
                    step=1,
                    key=f"novo_qtd_{i}"
                )
            if cat_selecionada and qtd_categoria > 0:
                detalhes_form.append((cat_selecionada, int(qtd_categoria)))
                qtd_somada += int(qtd_categoria)

        obs = st.text_area("Observações Adicionais", placeholder="Informações complementares sobre este lote...")

        col_btn_sub, col_info_total = st.columns([1.2, 2])
        with col_info_total:
            if qtd_somada > 0:
                st.info(f"🔢 **Total:** {qtd_somada} cabeças para **{prod_lancamento}**")
            else:
                st.warning("⚠️ Informe a quantidade de pelo menos uma categoria.")

        with col_btn_sub:
            submitted = st.form_submit_button("💾 Salvar Lançamento", type="primary")

        if submitted:
            if qtd_somada <= 0:
                st.error("A quantidade total de animais deve ser maior que zero!")
            else:
                op_banco = "CRÉDITO" if "CRÉDITO" in operacao else "DÉBITO"
                add_movimentacao(
                    produtor=prod_lancamento,
                    tipo_lancamento=tipo_lanc,
                    data=data_mov,
                    operacao=op_banco,
                    qtd_total=qtd_somada,
                    gta=gta,
                    nfp=nfp,
                    detalhes=detalhes_form,
                    observacoes=obs
                )
                st.success(f"✅ Lançamento de {qtd_somada} cabeças salvo com sucesso para {prod_lancamento}!")
                st.rerun()

    col_add_rem1, col_add_rem2, _ = st.columns([1, 1, 2])
    with col_add_rem1:
        if st.button("➕ Adicionar Mais Uma Categoria no Lote"):
            st.session_state.num_linhas_categoria += 1
            st.rerun()
    with col_add_rem2:
        if st.session_state.num_linhas_categoria > 1:
            if st.button("➖ Remover Linha"):
                st.session_state.num_linhas_categoria -= 1
                st.rerun()

# ==========================================
# ABA 3: HISTÓRICO & EXTRATO
# ==========================================
with tab_extrato:
    st.subheader("📋 Histórico Completo de Lançamentos")
    st.caption("Consulte, filtre, edite ou exclua lançamentos de Marcelo e Michele.")

    with st.expander("🔍 Filtros de Busca e Pesquisa", expanded=True):
        f_col0, f_col1, f_col2, f_col3, f_col4 = st.columns(5)
        with f_col0:
            f_prod = st.selectbox(
                "Produtor", 
                ["TODOS"] + lista_produtores,
                index=0 if produtor_selecionado == "Todos os Produtores" else (lista_produtores.index(produtor_selecionado) + 1 if produtor_selecionado in lista_produtores else 0)
            )
        with f_col1:
            tipo_filtro = st.selectbox("Lançamento", ["TODOS"] + TIPOS_LANCAMENTO_PADRAO)
        with f_col2:
            op_filtro = st.selectbox("Operação", ["TODOS", "CRÉDITO", "DÉBITO"])
        with f_col3:
            todas_cats = ["TODAS"] + get_categorias_cadastradas()
            cat_filtro = st.selectbox("Categoria", todas_cats)
        with f_col4:
            termo_busca = st.text_input("GTA / NFP / Texto", placeholder="Ex: AD-616710")

        f_data1, f_data2, _ = st.columns([1, 1, 2])
        with f_data1:
            usar_filtro_data = st.checkbox("Filtrar por Período de Datas")
        with f_data2:
            if usar_filtro_data:
                d_ini = st.date_input("Data Inicial", value=datetime.date(2023, 1, 1), format="DD/MM/YYYY")
                d_fim = st.date_input("Data Final", value=datetime.date.today() + datetime.timedelta(days=365), format="DD/MM/YYYY")
            else:
                d_ini = None
                d_fim = None

    filtros_dict = {
        "produtor": f_prod,
        "tipo_lancamento": tipo_filtro,
        "operacao": op_filtro,
        "categoria": cat_filtro,
        "busca_texto": termo_busca,
        "data_inicio": d_ini if usar_filtro_data else None,
        "data_fim": d_fim if usar_filtro_data else None
    }

    df_tabela = get_todas_movimentacoes(filtros_dict)

    if not df_tabela.empty:
        col_info_tabela, col_btn_tabela = st.columns([3, 1])
        with col_info_tabela:
            st.write(f"Mostrando **{len(df_tabela)}** lançamentos encontrados.")
        with col_btn_tabela:
            excel_filtrado = gerar_relatorio_excel_bytes(filtros_dict)
            st.download_button(
                label="📥 Exportar Dados Filtrados",
                data=excel_filtrado,
                file_name="lancamentos_filtrados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        colunas_mostrar = ["id", "PRODUTOR", "LANÇAMENTO", "DATA", "OPERAÇÃO", "QTD. TOTAL", "GTA", "NFP", "ANIMAIS / MESES", "OBSERVAÇÕES"]
        st.dataframe(
            df_tabela[colunas_mostrar],
            column_config={
                "id": st.column_config.NumberColumn("ID", width="small"),
                "PRODUTOR": st.column_config.TextColumn("Produtor", width="small"),
                "LANÇAMENTO": st.column_config.TextColumn("Lançamento", width="medium"),
                "DATA": st.column_config.TextColumn("Data", width="small"),
                "OPERAÇÃO": st.column_config.TextColumn("Operação", width="small"),
                "QTD. TOTAL": st.column_config.NumberColumn("Qtd. Total", format="%d", width="small"),
                "GTA": st.column_config.TextColumn("GTA", width="small"),
                "NFP": st.column_config.TextColumn("NFP", width="small"),
                "ANIMAIS / MESES": st.column_config.TextColumn("Animais / Meses", width="large"),
                "OBSERVAÇÕES": st.column_config.TextColumn("Observações", width="medium"),
            },
            use_container_width=True,
            hide_index=True
        )

        st.divider()
        st.subheader("⚙️ Ações no Registro (Editar / Excluir)")
        
        col_sel_id, col_btn_edit, col_btn_del = st.columns([2, 1, 1])
        with col_sel_id:
            ids_disponiveis = df_tabela["id"].tolist()
            id_selecionado = st.selectbox("Selecione o ID do Lançamento:", ids_disponiveis)
        
        with col_btn_del:
            st.write("")
            st.write("")
            if st.button("🗑️ Excluir Registro", type="secondary"):
                delete_movimentacao(id_selecionado)
                st.success(f"Lançamento #{id_selecionado} excluído!")
                st.rerun()

        with col_btn_edit:
            st.write("")
            st.write("")
            editar_modo = st.checkbox("✏️ Editar Registro")

        if editar_modo and id_selecionado:
            mov_atual = get_movimentacao(id_selecionado)
            if mov_atual:
                st.markdown(f"#### Editando Lançamento #{id_selecionado}")
                with st.form(f"form_editar_{id_selecionado}"):
                    ce0, ce1, ce2, ce3 = st.columns(4)
                    with ce0:
                        idx_p = lista_produtores.index(mov_atual.get("produtor", "Michele")) if mov_atual.get("produtor") in lista_produtores else 0
                        edit_prod = st.selectbox("Produtor", lista_produtores, index=idx_p)
                    with ce1:
                        try:
                            data_obj = datetime.datetime.strptime(mov_atual["data"], "%Y-%m-%d").date()
                        except Exception:
                            data_obj = datetime.date.today()
                        edit_data = st.date_input("Data", value=data_obj, format="DD/MM/YYYY")
                    with ce2:
                        idx_tipo = TIPOS_LANCAMENTO_PADRAO.index(mov_atual["tipo_lancamento"]) if mov_atual["tipo_lancamento"] in TIPOS_LANCAMENTO_PADRAO else 0
                        edit_tipo = st.selectbox("Lançamento", TIPOS_LANCAMENTO_PADRAO, index=idx_tipo)
                    with ce3:
                        idx_op = 0 if mov_atual["operacao"] == "CRÉDITO" else 1
                        edit_op = st.selectbox("Operação", ["CRÉDITO", "DÉBITO"], index=idx_op)

                    ce4, ce5, ce6 = st.columns(3)
                    with ce4:
                        edit_qtd = st.number_input("Qtd Total", value=int(mov_atual["qtd_total"]), min_value=1)
                    with ce5:
                        edit_gta = st.text_input("GTA", value=mov_atual["gta"] or "")
                    with ce6:
                        edit_nfp = st.text_input("NFP", value=mov_atual["nfp"] or "")

                    edit_obs = st.text_area("Observações", value=mov_atual["observacoes"] or "")
                    
                    detalhes_atuais = [(d["categoria"], d["quantidade"]) for d in mov_atual.get("detalhes", [])]

                    salvar_edicao = st.form_submit_button("💾 Salvar Alterações", type="primary")
                    if salvar_edicao:
                        update_movimentacao(
                            mov_id=id_selecionado,
                            produtor=edit_prod,
                            tipo_lancamento=edit_tipo,
                            data=edit_data,
                            operacao=edit_op,
                            qtd_total=edit_qtd,
                            gta=edit_gta,
                            nfp=edit_nfp,
                            detalhes=detalhes_atuais if detalhes_atuais else [("DIVERSOS / OUTROS", edit_qtd)],
                            observacoes=edit_obs
                        )
                        st.success("Lançamento atualizado com sucesso!")
                        st.rerun()

    else:
        st.info("Nenhum lançamento encontrado com os filtros selecionados.")

# ==========================================
# ABA 4: IMPORTAR / EXPORTAR
# ==========================================
with tab_import_export:
    st.subheader("📤 Exportar e Importar Planilhas")

    # Exibe mensagem persistente de sucesso se houver
    if st.session_state.msg_sucesso_importacao:
        st.success(st.session_state.msg_sucesso_importacao)
        st.balloons()
        st.session_state.msg_sucesso_importacao = None

    col_exp, col_imp = st.columns(2)

    with col_exp:
        st.markdown("### 📥 Exportar para Excel (.xlsx)")
        st.write("Baixe a planilha completa com abas de Lançamentos, Saldo por Categoria e Resumo por Produtor.")
        
        prod_exp = st.selectbox("Exportar dados de:", ["Todos os Produtores"] + lista_produtores, key="exp_prod_sel")
        excel_bytes = gerar_relatorio_excel_bytes({"produtor": prod_exp})
        st.download_button(
            label=f"📊 Baixar Planilha ({prod_exp})",
            data=excel_bytes,
            file_name=f"lancamentos_gado_{prod_exp.lower().replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        st.write("")
        st.divider()
        st.markdown("### 🔄 Restaurar Dados Iniciais da Foto")
        st.write("Deseja restaurar os 19 lançamentos originais da Michele?")
        if st.button("⚠️ Restaurar Lançamentos Padrão da Foto", type="secondary"):
            popular_banco(forcar=True)
            st.success("Dados padrão da planilha restaurados com sucesso para Michele!")
            st.rerun()

    with col_imp:
        st.markdown("### 📂 Importar Planilha Excel / CSV")
        st.write("Envie um arquivo `.xlsx` ou `.csv` para carregar novos lançamentos.")
        
        prod_imp_padrao = st.selectbox("Atribuir lançamentos para:", lista_produtores, key="imp_prod_sel")
        uploaded_file = st.file_uploader("Selecione o arquivo Excel ou CSV", type=["xlsx", "xls", "csv"], key="uploader_excel")
        
        if uploaded_file is not None:
            try:
                df_preview = carregar_dataframe_inteligente(uploaded_file)
                st.markdown("##### 👁️ Pré-visualização dos Dados Encontrados:")
                st.dataframe(df_preview.head(5), use_container_width=True)
                st.caption(f"Total de linhas detectadas: **{len(df_preview)}**")

                if st.button("🚀 Confirmar e Importar no Banco de Dados", type="primary"):
                    uploaded_file.seek(0)
                    qtd_imp, erros = importar_planilha_excel(uploaded_file, produtor_padrao=prod_imp_padrao)
                    if qtd_imp > 0:
                        st.session_state.msg_sucesso_importacao = f"🎉 **{qtd_imp} lançamentos importados com sucesso para {prod_imp_padrao}!**"
                        st.rerun()
                    if erros:
                        st.warning(f"{len(erros)} avisos durante a importação:")
                        for err in erros[:5]:
                            st.write(f"- {err}")
            except Exception as e:
                st.error(f"❌ Erro ao ler a planilha: {str(e)}")

# ==========================================
# ABA 5: USUÁRIOS & CONFIGURAÇÕES
# ==========================================
with tab_ajustes:
    st.subheader("⚙️ Usuários, Produtores e Categorias")

    col_sec0, col_sec1, col_sec2 = st.columns(3)

    # Gestão de Usuários
    with col_sec0:
        st.markdown("### 👥 Usuários do Sistema")
        users = get_usuarios_cadastrados()
        df_u = pd.DataFrame(users)[["username", "nome_completo", "produtor_associado"]]
        df_u.columns = ["Usuário", "Nome", "Produtor Padrão"]
        st.dataframe(df_u, use_container_width=True, hide_index=True)

        with st.form("form_novo_usuario_cad"):
            st.markdown("##### Cadastrar Novo Usuário")
            new_u = st.text_input("Usuário (Login)", placeholder="Ex: joao")
            new_p = st.text_input("Senha", type="password", placeholder="Ex: senha123")
            new_nome = st.text_input("Nome Completo", placeholder="Ex: João da Silva")
            new_assoc = st.selectbox("Produtor Associado", ["Todos os Produtores"] + lista_produtores)
            
            btn_add_u = st.form_submit_button("➕ Criar Usuário", type="primary")
            if btn_add_u:
                if new_u.strip() and new_p.strip() and new_nome.strip():
                    if add_novo_usuario(new_u, new_p, new_nome, new_assoc):
                        st.success(f"Usuário '{new_u}' criado com sucesso!")
                        st.rerun()
                    else:
                        st.error("Nome de usuário já existe.")
                else:
                    st.warning("Preencha todos os campos do usuário.")

    # Gestão de Produtores
    with col_sec1:
        st.markdown("### 🌾 Produtores Rurais")
        df_p = pd.DataFrame({"Produtor Rural": lista_produtores})
        st.dataframe(df_p, use_container_width=True, hide_index=True)

        with st.form("form_novo_produtor"):
            st.markdown("##### Cadastrar Novo Produtor")
            novo_prod_nome = st.text_input("Nome do Produtor", placeholder="Ex: Fazenda Santa Maria")
            btn_add_p = st.form_submit_button("➕ Adicionar Produtor", type="primary")
            if btn_add_p:
                if novo_prod_nome.strip():
                    if add_produtor(novo_prod_nome):
                        st.success(f"Produtor '{novo_prod_nome.strip()}' adicionado!")
                        st.rerun()
                    else:
                        st.error("Este produtor já existe.")
                else:
                    st.warning("Informe o nome do produtor.")

    # Gestão de Categorias
    with col_sec2:
        st.markdown("### 🏷️ Categorias de Animais")
        cats = get_categorias_cadastradas()
        df_cats = pd.DataFrame({"Categoria Cadastrada": cats})
        st.dataframe(df_cats, use_container_width=True, hide_index=True)

        with st.form("form_nova_categoria"):
            st.markdown("##### Cadastrar Nova Categoria")
            nova_cat = st.text_input("Nome da Categoria", placeholder="Ex: BEZERROS DESMAMADOS")
            faixa = st.text_input("Faixa Etária", placeholder="Ex: 8 a 12 meses")
            sexo = st.selectbox("Sexo", ["Misto", "Macho", "Fêmea"])
            
            btn_add_cat = st.form_submit_button("➕ Adicionar Categoria", type="primary")
            if btn_add_cat:
                if nova_cat.strip():
                    if add_categoria_personalizada(nova_cat, faixa, sexo):
                        st.success(f"Categoria '{nova_cat.upper()}' cadastrada!")
                        st.rerun()
                    else:
                        st.error("Esta categoria já existe.")
                else:
                    st.warning("Informe o nome da categoria.")

# Rodapé
st.markdown("---")
st.markdown("<div style='text-align: center; color: #888; font-size: 12px;'>Sistema de Gestão de Rebanho Bovino • Controle Seguro de Marcelo & Michele</div>", unsafe_allow_html=True)
