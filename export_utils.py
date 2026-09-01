import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import (
    get_todas_movimentacoes, get_saldo_por_categoria, get_resumo_totais, 
    get_resumo_por_produtores, add_movimentacao
)
import re

def gerar_relatorio_excel_bytes(filtros=None):
    """
    Gera uma planilha Excel estilizada e profissional com múltiplas abas:
    1. Lançamentos (com coluna Produtor)
    2. Saldo por Categoria
    3. Resumo por Produtor e Geral
    """
    output = io.BytesIO()
    
    produtor_filtro = filtros.get("produtor") if filtros else None
    df_mov = get_todas_movimentacoes(filtros)
    df_saldo = get_saldo_por_categoria(produtor_filtro)
    resumo = get_resumo_totais(produtor_filtro)
    df_resumo_prods = get_resumo_por_produtores()

    # Prepara DataFrame de Lançamentos para exportação
    if not df_mov.empty:
        colunas_export = ["PRODUTOR", "LANÇAMENTO", "DATA", "OPERAÇÃO", "QTD. TOTAL", "GTA", "NFP", "ANIMAIS / MESES", "OBSERVAÇÕES"]
        df_export = df_mov[colunas_export].copy()
    else:
        df_export = pd.DataFrame(columns=["PRODUTOR", "LANÇAMENTO", "DATA", "OPERAÇÃO", "QTD. TOTAL", "GTA", "NFP", "ANIMAIS / MESES", "OBSERVAÇÕES"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Lançamentos", index=False)
        df_saldo.to_excel(writer, sheet_name="Saldo por Categoria", index=False)
        df_resumo_prods.to_excel(writer, sheet_name="Resumo por Produtor", index=False)

        # Aba Resumo Consolidado
        prod_label = produtor_filtro if produtor_filtro and produtor_filtro not in ("TODOS", "Todos os Produtores") else "Consolidado (Todos)"
        df_resumo = pd.DataFrame([
            {"Indicador": "Produtor Selecionado", "Valor": prod_label},
            {"Indicador": "Saldo Atual do Rebanho (Cabeças)", "Valor": resumo["saldo_atual"]},
            {"Indicador": "Total de Entradas (Créditos)", "Valor": resumo["total_entradas"]},
            {"Indicador": "Total de Saídas (Débitos)", "Valor": resumo["total_saidas"]},
            {"Indicador": "Total de Movimentações Registradas", "Valor": resumo["total_registros"]}
        ])
        df_resumo.to_excel(writer, sheet_name="Resumo Geral", index=False)

        # Estilização das abas com openpyxl
        workbook = writer.book
        
        # Estilos
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        data_font = Font(name="Arial", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for sheet in workbook.worksheets:
            sheet.views.sheetView[0].showGridLines = True
            
            # Formata cabeçalho
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formata células de dados
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0"
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Ajusta largura das colunas
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output.seek(0)
    return output.getvalue()

def extrair_detalhes_de_texto(texto_categorias, qtd_total=0):
    if not texto_categorias or pd.isna(texto_categorias) or not str(texto_categorias).strip():
        return [("DIVERSOS / OUTROS", qtd_total)]
    
    texto = str(texto_categorias).strip()
    padrao = r'(\d+)\s*[-–]\s*([A-Za-zÀ-ÿ0-9\+\s]+?)(?=(?:\s+\d+\s*[-–]|$|\|))'
    matches = re.findall(padrao, texto)
    
    if matches:
        detalhes = []
        for qtd_str, cat_str in matches:
            cat_limpa = cat_str.strip().upper()
            qtd_num = int(qtd_str)
            detalhes.append((cat_limpa, qtd_num))
        return detalhes
    else:
        return [(texto.upper(), qtd_total)]

def importar_planilha_excel(arquivo_bytes_ou_upload, produtor_padrao="Michele"):
    """
    Lê uma planilha Excel/CSV e importa os lançamentos para o banco de dados.
    """
    try:
        df = pd.read_excel(arquivo_bytes_ou_upload)
    except Exception:
        try:
            arquivo_bytes_ou_upload.seek(0)
            df = pd.read_csv(arquivo_bytes_ou_upload)
        except Exception as e:
            return 0, [f"Erro ao abrir arquivo: {str(e)}"]

    col_map = {}
    for col in df.columns:
        c_upper = str(col).strip().upper()
        if "PROD" in c_upper:
            col_map[col] = "PRODUTOR"
        elif "LANÇ" in c_upper or "LANC" in c_upper or "TIPO" in c_upper:
            col_map[col] = "LANÇAMENTO"
        elif "DATA" in c_upper or "DARA" in c_upper:
            col_map[col] = "DATA"
        elif "OPER" in c_upper:
            col_map[col] = "OPERAÇÃO"
        elif "QTD" in c_upper or "QUANT" in c_upper or "TOTAL" in c_upper:
            col_map[col] = "QTD. TOTAL"
        elif "GTA" in c_upper:
            col_map[col] = "GTA"
        elif "NFP" in c_upper or "NOTA" in c_upper or "NF" in c_upper:
            col_map[col] = "NFP"
        elif "ANIMA" in c_upper or "CATEG" in c_upper or "MESES" in c_upper:
            col_map[col] = "ANIMAIS / MESES"
        elif "OBS" in c_upper:
            col_map[col] = "OBSERVAÇÕES"

    df = df.rename(columns=col_map)
    
    campos_obrigatorios = ["LANÇAMENTO", "DATA", "OPERAÇÃO", "QTD. TOTAL"]
    for campo in campos_obrigatorios:
        if campo not in df.columns:
            return 0, [f"Coluna obrigatória não encontrada: '{campo}'. Verifique o cabeçalho da sua planilha."]

    total_importados = 0
    erros = []

    for index, row in df.iterrows():
        try:
            tipo_lancamento = str(row["LANÇAMENTO"]).strip()
            if pd.isna(tipo_lancamento) or not tipo_lancamento or tipo_lancamento == "nan":
                continue

            produtor = str(row["PRODUTOR"]).strip() if "PRODUTOR" in df.columns and pd.notna(row["PRODUTOR"]) and str(row["PRODUTOR"]) != "nan" else produtor_padrao

            raw_data = row["DATA"]
            if isinstance(raw_data, (pd.Timestamp, pd.DatetimeIndex)):
                data_str = raw_data.strftime("%d/%m/%Y")
            elif hasattr(raw_data, "strftime"):
                data_str = raw_data.strftime("%d/%m/%Y")
            else:
                data_str = str(raw_data).strip()

            operacao = str(row["OPERAÇÃO"]).strip().upper()
            if "CR" in operacao or "ENTRADA" in operacao or "+" in operacao:
                operacao = "CRÉDITO"
            else:
                operacao = "DÉBITO"

            try:
                qtd_total = int(float(str(row["QTD. TOTAL"]).replace(".", "").replace(",", ".")))
            except Exception:
                qtd_total = 0

            gta = str(row["GTA"]).strip() if "GTA" in df.columns and pd.notna(row["GTA"]) and str(row["GTA"]) != "nan" else ""
            nfp = str(row["NFP"]).strip() if "NFP" in df.columns and pd.notna(row["NFP"]) and str(row["NFP"]) != "nan" else ""
            obs = str(row["OBSERVAÇÕES"]).strip() if "OBSERVAÇÕES" in df.columns and pd.notna(row["OBSERVAÇÕES"]) and str(row["OBSERVAÇÕES"]) != "nan" else ""
            animais_texto = str(row["ANIMAIS / MESES"]).strip() if "ANIMAIS / MESES" in df.columns and pd.notna(row["ANIMAIS / MESES"]) and str(row["ANIMAIS / MESES"]) != "nan" else ""

            detalhes = extrair_detalhes_de_texto(animais_texto, qtd_total)

            add_movimentacao(
                produtor=produtor,
                tipo_lancamento=tipo_lancamento,
                data=data_str,
                operacao=operacao,
                qtd_total=qtd_total,
                gta=gta,
                nfp=nfp,
                detalhes=detalhes,
                observacoes=obs
            )
            total_importados += 1
        except Exception as e:
            erros.append(f"Linha {index + 2}: {str(e)}")

    return total_importados, erros
